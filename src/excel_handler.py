"""
Author: Luigi Ferraioli
Copyright: © 2025 Luigi Ferraioli
"""

import os
import math
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side

from typing import Optional
from file_handler import IFileHandler
from openpyxl.utils import get_column_letter


class ExcelHandler(IFileHandler):
    def __init__(self, config: dict):
        self.config = config
        self.output_path = config.get("Speicherort", os.path.expanduser("~"))
        self.customer_data = {}

        if not os.path.exists(self.output_path):
            os.makedirs(self.output_path)

    def set_customer_data(self, data: dict):
        translated = {
            "Anrede": data.get("salutation", ""),
            "Nachname": data.get("last_name", ""),
            "Vorname": data.get("first_name", ""),
            "Straße": data.get("street", ""),
            "Hausnummer": data.get("number", ""),
            "PLZ": data.get("postal_code", ""),
            "Ort": data.get("city", "")
        }

        self.customer_data = translated

    def build_filename(self, customer_data: dict, base_name: str = "report") -> str:
        parts = []

        # Add name if requested
        if self.config.get("Name", False):
            last_name = customer_data.get("last_name", "")
            first_name = customer_data.get("first_name", "")
            name_part = f"{last_name}_{first_name}".strip("_")
            if name_part:
                parts.append(name_part)

        # Add date if requested
        if self.config.get("Datum", False):
            date_str = datetime.now().strftime("%Y-%m-%d")
            parts.append(date_str)

        # Add time if requested
        if self.config.get("Uhrzeit", False):
            time_str = datetime.now().strftime("%H-%M-%S")
            parts.append(time_str)

        if not parts:
            parts.append(base_name)

        filename = "_".join(parts) + ".xlsx"
        return filename

    def create_file(self, df: pd.DataFrame, filename: Optional[str] = "bericht.xlsx") -> None:
        file_path = os.path.join(self.output_path, filename)

        # Spaltenüberschriften anpassen
        df = self.append_units(df)

        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            if self.customer_data:
                customer_df = pd.DataFrame.from_dict(
                    self.customer_data, orient="index", columns=["Wert"])
                customer_df.index.name = "Feld"
                customer_df.to_excel(writer, sheet_name="Kundendaten")

            df.to_excel(writer, sheet_name="Daten", index=False)

        self._adjust_rows_and_columns(file_path, sheet_name="Kundendaten")
        self._adjust_rows_and_columns(file_path, sheet_name="Daten")

    def append_units(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Appends units to column headers for specific columns:
        - 'Länge', 'Breite', 'Höhe1', 'Höhe2' → adds unit (e.g., "Länge [m]")
        - 'Ergebnis' → adds squared unit (e.g., "Ergebnis [m²]")

        Args:
            df (pd.DataFrame): Input DataFrame

        Returns:
            pd.DataFrame: A copy of the input DataFrame with updated column headers.
        """
        df = df.copy()
        einheit = self.config.get("Einheit", "")

        new_columns = []
        for col in df.columns:
            if col in ["Länge", "Breite", "Höhe1", "Höhe2"]:
                new_columns.append(f"{col} [{einheit}]")
            elif col == "Ergebnis":
                new_columns.append(f"{col} [{einheit}²]")
            else:
                # Flächenart oder andere Spalten unverändert
                new_columns.append(col)

        df.columns = new_columns
        return df

    def load_file(self, file_path: str,
                  table_columns=None) -> tuple[Optional[dict], Optional[pd.DataFrame]]:
        """
        Loads an Excel file created by this handler.

        Args:
            file_path (str): Path to the Excel file.
            table_columns (list, optional): List of column names to assign to the data DataFrame.

        Returns:
            tuple: A tuple containing:
                - customer_data (dict or None): Dictionary of customer data if loaded successfully, else None.
                - data_df (pd.DataFrame or None): DataFrame with the main data if loaded successfully, else None.
        """
        try:
            customer_df = pd.read_excel(
                file_path, sheet_name="Kundendaten", index_col=0)
            customer_data = customer_df["Wert"].to_dict()

            data_df = pd.read_excel(file_path, sheet_name="Daten")

            if table_columns is not None:
                data_df.columns = table_columns

                if "Bemerkung" in data_df.columns:
                    data_df["Bemerkung"] = data_df["Bemerkung"].fillna("")

                float_cols = ["Länge", "Breite", "Höhe1", "Höhe2"]
                for col in float_cols:
                    if col in data_df.columns:
                        data_df[col] = data_df[col].astype(float)

            if table_columns is None:
                for column in data_df.columns:
                    data_df[column] = data_df[column].fillna("")

            return customer_data, data_df

        except Exception:
            return None, None

    def _adjust_rows_and_columns(self, file_path: str,
                                 sheet_name: str = "Daten", col_width: int = 60):
        """
        Reads data from an Excel file and prepares the DataFrame.

        - Loads the sheet named "Daten".
        - Replaces all NaN values with empty strings.
        - Optionally converts all integer columns to float for consistency.

        Parameters:
            file_path (str): The path to the Excel file.

        Returns:
            pandas.DataFrame: A cleaned DataFrame with missing values filled
            and numeric columns optionally converted to float.
        """
        wb = load_workbook(file_path)
        ws = wb[sheet_name]

        header_fill = PatternFill(
            start_color="A9C4EB", end_color="A9C4EB", fill_type="solid")
        alt_fill = PatternFill(start_color="F2F2F2",
                               end_color="F2F2F2", fill_type="solid")
        bold_font = Font(bold=True)
        border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000")
        )

        for row_idx, row in enumerate(ws.iter_rows(), start=1):
            max_lines = 1
            is_header = row_idx == 1

            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

                if is_header:
                    cell.fill = header_fill
                    cell.font = bold_font
                elif row_idx % 2 == 0:
                    cell.fill = alt_fill

                cell.border = border

                # Dynamische Höhenanpassung
                if cell.value:
                    max_lines = max(max_lines, math.ceil(
                        len(str(cell.value)) / col_width))

            ws.row_dimensions[row_idx].height = max_lines * 15

        # Spaltenbreiten
        for col_idx in range(1, ws.max_column + 1):
            max_length = max(len(str(ws.cell(row=row, column=col_idx).value or ""))
                             for row in range(1, ws.max_row + 1))
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = min(
                col_width, max_length + 2)

        wb.save(file_path)
